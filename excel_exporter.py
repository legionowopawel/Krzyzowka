# -*- coding: utf-8 -*-
"""
excel_exporter.py — Export krzyżówki do pliku Excel (.xlsx)
Formatowanie: idealne kwadraty, czarne ramki, pytania poniżej
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from typing import Optional
from crossword_grid import CrosswordGrid


class ExcelExporter:
    """Eksportuje krzyżówkę do Excel (.xlsx)."""
    
    def __init__(self):
        """Inicjalizacja."""
        # Styl dla komórki z literą
        self.border_black_thick = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Styl dla komórki czarnej (niedostępnej)
        self.fill_black = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        
        # Styl dla komórki pustej
        self.fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Font dla liter
        self.font_letter = Font(name='Arial', size=12, bold=True)
        
        # Font dla numerów pytań
        self.font_clue_num = Font(name='Arial', size=8, bold=True, color='FF0000')
        
        # Alignment
        self.align_center = Alignment(horizontal='center', vertical='center')
    
    def export(self, grid: CrosswordGrid, filepath: str) -> bool:
        """
        Eksportuj krzyżówkę do pliku Excel.
        
        Args:
            grid: Siatka krzyżówki
            filepath: Ścieżka do pliku wyjściowego
        
        Returns:
            True jeśli OK, False jeśli błąd
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Krzyżówka"
            
            # Ustaw rozmiar komórek (kwadraty)
            cell_size = 30  # pixel
            for col in range(1, grid.width + 1):
                ws.column_dimensions[get_column_letter(col)].width = 4.5
            for row in range(1, grid.height + 1):
                ws.row_dimensions[row].height = cell_size
            
            # Rysuj siatkę
            for r in range(grid.height):
                for c in range(grid.width):
                    row_excel = r + 1
                    col_excel = c + 1
                    col_letter = get_column_letter(col_excel)
                    
                    cell = ws[f'{col_letter}{row_excel}']
                    
                    grid_cell = grid.grid[r][c]
                    
                    if grid_cell is None:
                        # Czarna komórka (niedostępna)
                        cell.fill = self.fill_black
                        cell.border = Border()  # Bez ramki
                    else:
                        # Biała/beżowa komórka
                        cell.fill = self.fill_white
                        cell.border = self.border_black_thick
                        cell.alignment = self.align_center
                        
                        # Jeśli komórka zawiera literę, wstaw ją
                        if grid_cell and grid_cell != "":
                            cell.value = grid_cell
                            cell.font = self.font_letter
                        
                        # Jeśli komórka ma numer pytania, wstaw go
                        clue_num = grid.get_clue_number(r, c)
                        if clue_num is not None:
                            # Wstaw numer w lewym górnym rogu jako mały tekst
                            # (opcjonalnie: można to zrobić jako życzenie komentarzy)
                            # Na razie: dodaj jako bardzo mały tekst do wartości
                            if grid_cell:
                                # Nie nadpisuj literę — dodajmy numer do komentarza
                                comment_text = f"Pytanie: {clue_num}"
                                from openpyxl.comments import Comment
                                cell.comment = Comment(comment_text, "Generator")
            
            # Dodaj puste wiersze do separation
            start_row = grid.height + 3
            
            # Dodaj pytania
            ws[f'A{start_row}'] = "PYTANIA - POZIOMO"
            ws[f'A{start_row}'].font = Font(name='Arial', size=12, bold=True)
            
            h_clues, v_clues = grid.get_clues_list()
            
            row = start_row + 1
            for num, clue, word in h_clues:
                ws[f'A{row}'] = f"{num}. {clue} ({len(word)} liter)"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1
            
            row += 1
            ws[f'A{row}'] = "PYTANIA - PIONOWO"
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            
            row += 1
            for num, clue, word in v_clues:
                ws[f'A{row}'] = f"{num}. {clue} ({len(word)} liter)"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1
            
            # Zapisz
            wb.save(filepath)
            print(f"[ExcelExporter] Zapisano: {filepath}")
            return True
            
        except Exception as e:
            print(f"[ExcelExporter] BŁĄD: {e}")
            return False
